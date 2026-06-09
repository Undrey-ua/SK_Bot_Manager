"""Імпорт продажів з Sales.xlsx (колонки як у Google Sheets «Продажі»)."""
from __future__ import annotations

import argparse
import asyncio
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook
from sqlalchemy import select

from config.settings import get_settings
from database.models import Brand, Client, User
from database.repositories.brand import BrandRepository
from database.repositories.sale import SaleRepository
from database.session import create_engine, create_session_factory

SHEET_TELEGRAM = {
    "андрій": 535827585,
    "роман": 5009921383,
    "павло": 7770797356,
}

BRAND_ALIASES: dict[str, str] = {
    "solida": "IVC: Divino",
    "ivc solida": "IVC: Divino",
    "ivc: solida": "IVC: Divino",
    "tarkett express": "Tarkett: Express",
    "tarkett: express": "Tarkett: Express",
    "tarkett living": "Tarkett Living",
    "tarkett herringbone": "Tarkett Herringbone",
    "berryalloc smartline": "BerryAlloc: Smartline",
    "berryalloc: smartline": "BerryAlloc: Smartline",
    "big carmelita": "BIG: Carmelita",
    "big: carmelita": "BIG: Carmelita",
    "big pureloc40": "BIG: Pureloc40",
    "big: pureloc40": "BIG: Pureloc40",
    "big novocore legacy": "BIG: Novocore Legacy",
    "big: novocore legacy": "BIG: Novocore Legacy",
    "carmelita": "BIG: Carmelita",
    "pureloc40": "BIG: Pureloc40",
    "novocore legacy": "BIG: Novocore Legacy",
    "berryalloc new: novocore legacy": "BIG: Novocore Legacy",
    "berryalloc new novocore legacy": "BIG: Novocore Legacy",
    "xpertpro": "XpertPro",
    "ivc divino": "IVC: Divino",
    "ivc: divino": "IVC: Divino",
    "prisma": "Prisma",
    "ado": "ADO",
    "ivc: solida метал": "IVC: Divino",
    "ivc solida метал": "IVC: Divino",
}

COMPOUND_BRANDS: dict[str, list[str]] = {
    "berryalloc: carmelita pureloc40": ["BIG: Carmelita", "BIG: Pureloc40"],
}


def _norm(text: object) -> str:
    return " ".join(str(text or "").strip().split()).casefold()


def _parse_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        from datetime import timedelta
        base = date(1899, 12, 30)
        return base + timedelta(days=int(value))
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _resolve_brand_names(raw: str) -> list[str]:
    key = _norm(raw)
    if key in COMPOUND_BRANDS:
        return COMPOUND_BRANDS[key]
    parts = [p.strip() for p in raw.replace(" + ", "+").split("+") if p.strip()]
    return parts or [raw.strip()]


def _map_brand(name: str, brand_by_key: dict[str, Brand]) -> Brand | None:
    key = _norm(name)
    if key in BRAND_ALIASES:
        name = BRAND_ALIASES[key]
        key = _norm(name)
    for bkey, brand in brand_by_key.items():
        if bkey == key or bkey in key or key in bkey:
            return brand
    return brand_by_key.get(key)


def _find_client(
    clients: list[Client],
    manager_id: int,
    shop: str,
    city: str,
) -> Client | None:
    shop_n = _norm(shop)
    city_n = _norm(city) if city else ""
    matches = [c for c in clients if c.manager_id == manager_id and _norm(c.name) == shop_n]
    if len(matches) == 1:
        return matches[0]
    if city_n:
        for c in matches:
            if city_n in _norm(c.address) or (
                c.comment and city_n in _norm(c.comment)
            ):
                return c
    return matches[0] if matches else None


async def run_import(path: Path, *, dry_run: bool) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        print("Порожній файл")
        return 1

    settings = get_settings()
    engine = create_engine(settings.database_url)
    factory = create_session_factory(engine)
    created = 0
    errors: list[str] = []

    async with factory() as session:
        users = {u.telegram_id: u for u in (await session.execute(select(User))).scalars()}
        clients = list((await session.execute(select(Client))).scalars())
        brands = list((await session.execute(select(Brand))).scalars())
        brand_by_key = {_norm(b.name): b for b in brands}
        sale_repo = SaleRepository(session)

        for row_num, row in enumerate(rows[1:], start=2):
            if not row or not row[0]:
                continue
            sold_at = _parse_date(row[0])
            manager_name = str(row[1] or "").strip()
            shop = str(row[2] or "").strip()
            city = str(row[3] or "").strip() if len(row) > 3 else ""
            stand_raw = str(row[5] or "").strip() if len(row) > 5 else ""
            qty_raw = row[6] if len(row) > 6 else None
            comment = str(row[7]).strip() if len(row) > 7 and row[7] else None

            line = f"рядок {row_num}"
            if not sold_at or not shop or not stand_raw:
                errors.append(f"{line}: пропущено (дата/ТТ/стенд)")
                continue
            try:
                qty = Decimal(str(qty_raw).replace(",", "."))
            except (InvalidOperation, TypeError):
                errors.append(f"{line}: некоректна кількість")
                continue

            mgr_key = _norm(manager_name.split()[0] if manager_name else "")
            telegram = SHEET_TELEGRAM.get(mgr_key)
            if not telegram or telegram not in users:
                errors.append(f"{line}: менеджер «{manager_name}»")
                continue
            manager = users[telegram]
            client = _find_client(clients, manager.id, shop, city)
            if client is None:
                errors.append(f"{line}: ТТ «{shop}» не знайдено")
                continue

            brand_names = _resolve_brand_names(stand_raw)
            per_brand_qty = qty / len(brand_names)

            for bn in brand_names:
                brand = _map_brand(bn, brand_by_key)
                if brand is None:
                    errors.append(f"{line}: бренд «{bn}»")
                    break
                if dry_run:
                    print(
                        f"dry-run {line}: {shop} → {brand.name} "
                        f"{per_brand_qty} м² ({sold_at})"
                    )
                else:
                    await sale_repo.create(
                        manager_id=manager.id,
                        client_id=client.id,
                        brand_id=brand.id,
                        quantity=per_brand_qty,
                        sold_at=sold_at,
                        comment=comment,
                    )
                created += 1
            else:
                continue

        if not dry_run:
            await session.commit()

    await engine.dispose()
    print(f"\nГотово: записів {created}, помилок {len(errors)}")
    for e in errors[:30]:
        print(f"  • {e}")
    return 0 if not errors else 1


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path, default=Path("Sales.xlsx"))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    sys.exit(asyncio.run(run_import(args.xlsx, dry_run=args.dry_run)))


if __name__ == "__main__":
    main()
