"""Привʼязка стендів до клієнтів з матриці Excel (число = к-сть встановлених).

Файл: Стенди.xlsx — вкладки «Андрій», «Роман», «Павло».
Колонки: Назва ТТ, Область, Місто, Адреса + назви стендів.

Запуск:
  .venv/bin/python scripts/import_stands_xlsx.py Стенди.xlsx --dry-run
  .venv/bin/python scripts/import_stands_xlsx.py Стенди.xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select

from config.settings import get_settings
from database.models import Client, ClientStand, Stand, User
from database.repositories.client import ClientRepository
from database.repositories.stand import StandRepository
from database.session import create_engine, create_session_factory

SHEET_TELEGRAM: dict[str, int] = {
    "андрій": 535827585,
    "роман": 5009921383,
    "павло": 7770797356,
}

STAND_ALIASES: dict[str, str] = {}


def _norm(text: object) -> str:
    return " ".join(str(text or "").strip().split()).casefold()


def _parse_quantity(value: object) -> int:
    """Як numberOrZero у SK_Account: 0/порожньо → 0, інакше ціла кількість."""
    if value is None:
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return max(0, int(float(value)))
    s = str(value).strip().lower().replace(",", ".")
    if s in ("", "0", "0.0", "none", "-"):
        return 0
    try:
        return max(0, int(float(s)))
    except ValueError:
        return 0


def _parse_sheet(ws: Worksheet) -> tuple[list[str], list[dict]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    header = rows[0]
    stand_names: list[str] = []
    stand_start = 4
    for cell in header[stand_start:]:
        name = str(cell).strip() if cell is not None else ""
        if name:
            stand_names.append(name)

    records: list[dict] = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or not any(row[:4]):
            continue
        name = str(row[0] or "").strip()
        if not name:
            continue
        stands: dict[str, int] = {}
        for i, stand_name in enumerate(stand_names):
            col_idx = stand_start + i
            val = row[col_idx] if col_idx < len(row) else None
            stands[stand_name] = _parse_quantity(val)
        records.append(
            {
                "row": str(row_num),
                "name": name,
                "address": str(row[3] or "").strip(),
                "stands": stands,
            }
        )
    return stand_names, records


def _read_workbook(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    result: list[dict] = []
    for sheet_name in wb.sheetnames:
        key = _norm(sheet_name)
        if key not in SHEET_TELEGRAM:
            continue
        stand_names, records = _parse_sheet(wb[sheet_name])
        for rec in records:
            rec["sheet"] = sheet_name
            rec["manager_key"] = key
            rec["_stand_columns"] = stand_names
        result.extend(records)
    wb.close()
    return result


def _index_clients(clients: list[Client]) -> tuple[
    dict[tuple[int, str, str], list[Client]],
    dict[tuple[int, str], list[Client]],
]:
    by_name_addr: dict[tuple[int, str, str], list[Client]] = {}
    by_name: dict[tuple[int, str], list[Client]] = {}
    for c in clients:
        k1 = (c.manager_id, _norm(c.name), _norm(c.address))
        by_name_addr.setdefault(k1, []).append(c)
        k2 = (c.manager_id, _norm(c.name))
        by_name.setdefault(k2, []).append(c)
    return by_name_addr, by_name


def _build_global_name_addr_index(
    by_name_addr: dict[tuple[int, str, str], list[Client]],
) -> dict[tuple[str, str], list[Client]]:
    global_index: dict[tuple[str, str], list[Client]] = {}
    for (_mid, name, addr), clients in by_name_addr.items():
        key = (name, addr)
        global_index.setdefault(key, []).extend(clients)
    return global_index


def _find_clients(
    *,
    manager_id: int,
    name: str,
    address: str,
    by_name_addr: dict[tuple[int, str, str], list[Client]],
    by_name: dict[tuple[int, str], list[Client]],
    global_by_name_addr: dict[tuple[str, str], list[Client]],
) -> list[Client]:
    hits = by_name_addr.get((manager_id, _norm(name), _norm(address)))
    if hits:
        return hits
    hits = by_name.get((manager_id, _norm(name)), [])
    if len(hits) == 1:
        return hits
    if len(hits) > 1 and address:
        addr_n = _norm(address)
        filtered = [c for c in hits if _norm(c.address) == addr_n]
        if filtered:
            return filtered
        filtered = [c for c in hits if addr_n in _norm(c.address) or _norm(c.address) in addr_n]
        if len(filtered) == 1:
            return filtered

    return global_by_name_addr.get((_norm(name), _norm(address)), [])


async def run_import(path: Path, *, dry_run: bool) -> int:
    records = _read_workbook(path)
    if not records:
        print("Немає даних для імпорту.")
        return 1

    settings = get_settings()
    engine = create_engine(settings.database_url)
    factory = create_session_factory(engine)

    updated_clients = 0
    links_added = 0
    errors: list[str] = []
    warnings: list[str] = []
    pending_stands: dict[int, dict[int, int]] = {}

    async with factory() as session:
        users = {u.telegram_id: u for u in (await session.execute(select(User))).scalars()}
        manager_by_key = {
            k: users[tg] for k, tg in SHEET_TELEGRAM.items() if tg in users
        }

        stand_repo = StandRepository(session)
        all_stands = list((await session.execute(select(Stand))).scalars())
        stand_by_name = {s.name: s for s in all_stands}
        max_sort = max((s.sort_order for s in all_stands), default=0)

        header_names = set()
        for rec in records:
            header_names.update(rec.get("_stand_columns") or rec["stands"].keys())

        for raw_name in sorted(header_names):
            canon = STAND_ALIASES.get(raw_name, raw_name)
            if canon not in stand_by_name:
                if dry_run:
                    warnings.append(f"буде створено стенд: {canon}")
                    stand = Stand(name=canon, sort_order=max_sort + 1, is_active=True)
                    stand.id = -(max_sort + 1)
                    max_sort += 1
                    stand_by_name[canon] = stand
                else:
                    stand = await stand_repo.create(canon, sort_order=max_sort + 1)
                    max_sort += 1
                    stand_by_name[canon] = stand
                    print(f"created stand: {canon}")

        clients = list((await session.execute(select(Client))).scalars())
        by_name_addr, by_name = _index_clients(clients)
        global_by_name_addr = _build_global_name_addr_index(by_name_addr)
        client_repo = ClientRepository(session)

        for rec in records:
            line = f"«{rec['sheet']}» рядок {rec['row']}"
            manager = manager_by_key.get(rec["manager_key"])
            if manager is None:
                errors.append(f"{line}: менеджера не знайдено")
                continue

            matched = _find_clients(
                manager_id=manager.id,
                name=rec["name"],
                address=rec["address"],
                by_name_addr=by_name_addr,
                by_name=by_name,
                global_by_name_addr=global_by_name_addr,
            )
            if not matched:
                errors.append(f"{line}: клієнта «{rec['name']}» не знайдено")
                continue
            if len(matched) > 1:
                warnings.append(
                    f"{line}: «{rec['name']}» — {len(matched)} записів у БД, "
                    "оновлюються всі"
                )

            stand_qty: dict[int, int] = {}
            stand_labels: list[str] = []
            row_error = False
            for stand_name, qty in rec["stands"].items():
                if qty <= 0:
                    continue
                canon = STAND_ALIASES.get(stand_name, stand_name)
                stand = stand_by_name.get(canon)
                if stand is None:
                    errors.append(f"{line}: невідомий стенд «{stand_name}»")
                    row_error = True
                    break
                stand_labels.append(f"{canon}×{qty}")
                if stand.id > 0:
                    stand_qty[stand.id] = stand_qty.get(stand.id, 0) + qty
            if row_error:
                continue

            for client in matched:
                if dry_run:
                    print(
                        f"dry-run {line}: {client.name} (id={client.id}) → "
                        f"{', '.join(stand_labels) or '—'}"
                    )
                    updated_clients += 1
                else:
                    bucket = pending_stands.setdefault(client.id, {})
                    for sid, q in stand_qty.items():
                        bucket[sid] = bucket.get(sid, 0) + q

        if not dry_run:
            clients_by_id = {c.id: c for c in clients}
            for client_id, qty_map in pending_stands.items():
                client = clients_by_id.get(client_id)
                if client is None:
                    continue
                stand_list = sorted(qty_map.keys())
                await client_repo.update(
                    client_id=client.id,
                    region_id=client.region_id,
                    name=client.name,
                    address=client.address,
                    comment=client.comment,
                    stand_ids=stand_list,
                    stand_qty=qty_map,
                )
                links_added += sum(qty_map.values())
                updated_clients += 1
            await session.commit()

    await engine.dispose()

    print()
    print(
        f"Готово: оновлено записів клієнтів {updated_clients}, "
        f"звʼязків стендів {links_added if not dry_run else '(dry-run)'}, "
        f"помилок {len(errors)}, попереджень {len(warnings)}"
    )
    for msg in warnings[:20]:
        print(f"  ! {msg}")
    if errors:
        print("\nПомилки:")
        for msg in errors[:40]:
            print(f"  • {msg}")
        if len(errors) > 40:
            print(f"  … ще {len(errors) - 40}")
    return 0 if not errors else 1


def main() -> None:
    parser = argparse.ArgumentParser(description="Імпорт стендів з матриці .xlsx")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    if not args.xlsx.is_file():
        print(f"Файл не знайдено: {args.xlsx}", file=sys.stderr)
        sys.exit(1)
    sys.exit(asyncio.run(run_import(args.xlsx, dry_run=args.dry_run)))


if __name__ == "__main__":
    main()
