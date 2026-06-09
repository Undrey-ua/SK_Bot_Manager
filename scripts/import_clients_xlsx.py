"""Імпорт клієнтів з Excel у БД.

Формат колонок на кожній вкладці:
  Назва ТТ | Область | Місто | Адреса

Підтримка:
  • окремі вкладки «Роман» і «Павло» (менеджер з назви аркуша);
  • один аркуш — менеджер визначається за колонкою «Область».

Запуск:
  .venv/bin/python scripts/import_clients_xlsx.py Clients.xlsx --dry-run
  .venv/bin/python scripts/import_clients_xlsx.py Clients.xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select

from config.settings import get_settings
from database.models import Client, ManagerRegion, User
from database.repositories.client import ClientRepository
from database.session import create_engine, create_session_factory

HEADER_ALIASES: dict[str, str] = {
    "назва тт": "name",
    "назва": "name",
    "назва точки": "name",
    "тт": "name",
    "область": "region",
    "регіон": "region",
    "місто": "city",
    "город": "city",
    "адреса": "address",
    "адрес": "address",
}

# Ключ аркуша (нижній регістр) → telegram_id менеджера
SHEET_TELEGRAM: dict[str, int] = {
    "роман": 5009921383,
    "павло": 7770797356,
    "андрій": 535827585,
}

# Області, які можуть бути лише у цього менеджера (інші з аркуша → інший менеджер)
MANAGER_PRIMARY_REGIONS: dict[int, frozenset[str]] = {
    535827585: frozenset({"Київ", "Київська", "Черкаська", "Житомирська"}),
}

REGION_ALIASES: dict[str, str] = {
    "днепропетровская": "Дніпропетровська",
    "дніпропетровская": "Дніпропетровська",
    "кировоградская": "Кіровоградська",
    "кіровоградская": "Кіровоградська",
    "запорожская": "Запорізька",
    "одесская": "Одеська",
    "полтавская": "Полтавська",
    "харьковская": "Харківська",
    "николаевская": "Миколаївська",
    "миколаївская": "Миколаївська",
    "винницкая": "Вінницька",
    "львовская": "Львівська",
    "ровенская": "Рівненська",
    "волынская": "Волинська",
    "тернопольская": "Тернопільська",
    "черновицкая": "Чернівецька",
    "хмельницкая": "Хмельницька",
    "закарпатская": "Закарпатська",
    "ивано-франковская": "Івано-Франківська",
    "івано-франковская": "Івано-Франківська",
    "івано франківська": "Івано-Франківська",
    "ивано франковская": "Івано-Франківська",
}


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_key(text: str) -> str:
    return " ".join(text.strip().lower().split())


def _normalize_region_name(name: str) -> str:
    key = _norm_key(name)
    if key in REGION_ALIASES:
        return REGION_ALIASES[key]
    return " ".join(name.strip().split())


def _parse_header_row(row: tuple) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        key = HEADER_ALIASES.get(_norm_key(_cell_str(cell)))
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _read_sheet_rows(
    ws: Worksheet,
    *,
    sheet_name: str,
    manager_key: str | None,
) -> list[dict[str, str]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    col_map = _parse_header_row(header_row)
    for required in ("name", "region", "address"):
        if required not in col_map:
            missing = {
                "name": "Назва ТТ",
                "region": "Область",
                "address": "Адреса",
            }[required]
            raise ValueError(
                f"Вкладка «{sheet_name}»: не знайдено колонку «{missing}»"
            )

    result: list[dict[str, str]] = []
    for row_num, row in enumerate(rows_iter, start=2):
        if not row or all(_cell_str(c) == "" for c in row):
            continue
        record = {
            "sheet": sheet_name,
            "manager_key": manager_key or "",
            "row": str(row_num),
            "name": _cell_str(row[col_map["name"]]),
            "region": _normalize_region_name(_cell_str(row[col_map["region"]])),
            "address": _cell_str(row[col_map["address"]]),
            "city": _cell_str(row[col_map["city"]]) if "city" in col_map else "",
        }
        if not record["name"] and not record["region"]:
            continue
        result.append(record)
    return result


def _manager_sheets(wb) -> list[tuple[str, str]]:
    matched: list[tuple[str, str]] = []
    for name in wb.sheetnames:
        key = _norm_key(name)
        if key in SHEET_TELEGRAM:
            matched.append((name, key))
    return matched


def _read_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    manager_sheets = _manager_sheets(wb)

    if manager_sheets:
        result: list[dict[str, str]] = []
        for sheet_name, manager_key in manager_sheets:
            result.extend(
                _read_sheet_rows(
                    wb[sheet_name],
                    sheet_name=sheet_name,
                    manager_key=manager_key,
                )
            )
        wb.close()
        return result

    ws = wb.active
    rows = _read_sheet_rows(ws, sheet_name=ws.title, manager_key=None)
    wb.close()
    return rows


def _build_address_city(city: str, address: str) -> tuple[str, str | None, str | None]:
    address = address.strip()
    city = city.strip() or None
    if not address and city:
        return city, city, None
    if not address:
        return "", None, None
    return address, city, None


def _line_label(rec: dict[str, str]) -> str:
    if rec.get("sheet"):
        return f"«{rec['sheet']}» рядок {rec['row']}"
    return f"рядок {rec['row']}"


def _resolve_region(
    region_name: str,
    sheet_manager: User,
    region_by_manager: dict[tuple[int, str], ManagerRegion],
) -> ManagerRegion | None:
    """Область менеджера вкладки; якщо не його — власник з іншого менеджера."""
    norm = _norm_key(_normalize_region_name(region_name))
    own = region_by_manager.get((sheet_manager.id, norm))
    if own is not None:
        return own

    primary = MANAGER_PRIMARY_REGIONS.get(sheet_manager.telegram_id)
    if primary is None or _normalize_region_name(region_name) in primary:
        return None

    for (manager_id, key), region in region_by_manager.items():
        if key == norm:
            return region
    return None


async def run_import(path: Path, *, dry_run: bool) -> int:
    records = _read_rows(path)
    if not records:
        print("Немає рядків для імпорту.")
        return 1

    settings = get_settings()
    engine = create_engine(settings.database_url)
    factory = create_session_factory(engine)

    created = 0
    skipped = 0
    errors: list[str] = []

    async with factory() as session:
        users = list((await session.execute(select(User))).scalars().all())
        manager_by_telegram = {u.telegram_id: u for u in users}
        manager_by_sheet_key: dict[str, User] = {}
        for sheet_key, telegram_id in SHEET_TELEGRAM.items():
            user = manager_by_telegram.get(telegram_id)
            if user:
                manager_by_sheet_key[sheet_key] = user

        region_result = await session.execute(
            select(ManagerRegion).order_by(ManagerRegion.manager_id, ManagerRegion.name)
        )
        regions = list(region_result.scalars().all())
        region_by_manager: dict[tuple[int, str], ManagerRegion] = {}
        region_by_name: dict[str, ManagerRegion] = {}
        for r in regions:
            region_by_manager[(r.manager_id, _norm_key(r.name))] = r
            region_by_name[_norm_key(r.name)] = r

        client_repo = ClientRepository(session)
        existing_keys: set[tuple[int, str]] = set()
        for row in await session.execute(select(Client.manager_id, Client.name)):
            manager_id, name = row
            existing_keys.add((manager_id, name.strip().casefold()))

        for rec in records:
            line = _line_label(rec)
            if not rec["name"]:
                errors.append(f"{line}: порожня назва ТТ")
                continue
            if not rec["region"]:
                errors.append(f"{line}: порожня область")
                continue

            manager_key = rec.get("manager_key") or ""
            if manager_key:
                sheet_manager = manager_by_sheet_key.get(manager_key)
                if sheet_manager is None:
                    errors.append(f"{line}: менеджера для вкладки не знайдено в БД")
                    continue
                region = _resolve_region(
                    rec["region"],
                    sheet_manager,
                    region_by_manager,
                )
                if region is None:
                    errors.append(
                        f"{line}: область «{rec['region']}» не знайдено в БД"
                    )
                    continue
                manager = next(
                    (u for u in users if u.id == region.manager_id),
                    None,
                )
            else:
                region = region_by_name.get(_norm_key(rec["region"]))
                if region is None:
                    errors.append(
                        f"{line}: невідома область «{rec['region']}»"
                    )
                    continue
                manager = next(
                    (u for u in users if u.id == region.manager_id),
                    None,
                )

            address, city, comment = _build_address_city(rec["city"], rec["address"])
            if not address:
                errors.append(f"{line}: порожня адреса")
                continue

            dup_key = (region.manager_id, rec["name"].casefold())
            if dup_key in existing_keys:
                skipped += 1
                print(f"skip {line}: вже є «{rec['name']}»")
                continue

            manager_label = manager.name if manager else str(region.manager_id)
            if dry_run:
                print(
                    f"dry-run {line}: {rec['name']} → {manager_label}, "
                    f"{region.name}, {address}"
                )
                created += 1
                continue

            await client_repo.create(
                manager_id=region.manager_id,
                region_id=region.id,
                name=rec["name"],
                address=address,
                city=city,
                comment=comment,
                stand_ids=[],
            )
            existing_keys.add(dup_key)
            created += 1
            print(f"ok {line}: {rec['name']} ({region.name})")

        if not dry_run:
            await session.commit()

    await engine.dispose()

    print()
    print(f"Готово: +{created}, пропущено дублікатів {skipped}, помилок {len(errors)}")
    if errors:
        print("\nПомилки:")
        for msg in errors[:50]:
            print(f"  • {msg}")
        if len(errors) > 50:
            print(f"  … ще {len(errors) - 50}")
    return 0 if not errors else 1


def main() -> None:
    parser = argparse.ArgumentParser(description="Імпорт клієнтів з .xlsx")
    parser.add_argument("xlsx", type=Path, help="Шлях до файлу Excel")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Лише перевірка, без запису в БД",
    )
    args = parser.parse_args()
    if not args.xlsx.is_file():
        print(f"Файл не знайдено: {args.xlsx}", file=sys.stderr)
        sys.exit(1)
    sys.exit(asyncio.run(run_import(args.xlsx, dry_run=args.dry_run)))


if __name__ == "__main__":
    main()
